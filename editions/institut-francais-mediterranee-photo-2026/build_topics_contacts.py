"""
INLAND Tunisia — Topics + Contacts spreadsheet builder.
Two sheets (Topics, Contacts) per _inland/deliverables.md. INLAND house style.
Run from this directory: `python3 build_topics_contacts.py`
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# House palette
BG = "F9F4EE"
ACCENT = "400505"
HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
ALT_FILL = PatternFill("solid", fgColor=BG)
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10, color="111111")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D8D2CA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOPICS_HEADERS = [
    "Topic", "Description", "Category", "Links", "Contacts",
    "Interesting for", "Ideas for development", "Notes",
]

TOPICS = [
    [
        "Charfia palm-fence fishing (Kerkennah)",
        "On the low, palm-fringed Kerkennah Islands off Sfax, fishermen plant date-palm fronds "
        "in rows across the shallows to build the charfia — a great V of woven palm that channels "
        "fish on the ebb tide into capture chambers, the grounds shared by a draw of lots and "
        "worked to a maritime calendar. UNESCO inscribed it in 2020. Sea-level rise, plastic and "
        "youth emigration now threaten a fishing architecture built entirely from the palm.",
        "Nature",
        "UNESCO ICH 2020 (charfia-fishing-in-the-kerkennah-islands-01566); Earth Journalism Network; New Arab",
        "Kerkennah fishing community; Sfax INSTM marine-heritage researchers; APAL (coastal agency)",
        "The defining cover image; the palm-fence standing in the sea; hands, tide and lamplight",
        "Follow one charfia family through the autumn setting of the fronds and a dawn lift of the traps",
        "Endangered; identify a named skipper/family on the ground; seasonal (autumn set)",
    ],
    [
        "Sejnane women’s pottery",
        "In the Mogods hills behind Bizerte, the women of Sejnane hand-build pots, animal "
        "figurines and the famous Sejnane dolls from wadi clay with no wheel and no kiln, firing "
        "in open ground and painting with two-tone geometric signs from tattoo and Berber "
        "weaving. The methods have protohistoric roots, passed mother to daughter for some three "
        "thousand years; every stage is performed by women. UNESCO-inscribed 2018.",
        "Arts & Culture",
        "UNESCO ICH 2018 (pottery-skills-of-the-women-of-sejnane-01406); The Arab Weekly; phys.org",
        "Sejnane women’s pottery associations/cooperatives; ONAT (national crafts office)",
        "Women-as-archive; ceramics older than the wheel; the open ground-firing; roadside sale",
        "Portrait one named potter through a full clay-to-firing cycle; the dolls and their meaning",
        "Living but market-pressed; name a maker with the local association",
    ],
    [
        "Djerba — Ibadi mosques & El Ghriba",
        "Djerba is the only place in the Arab world where Ibadi, Maliki, Hanafi, Christian and "
        "Jewish communities have shared one small, waterless island for centuries. Ibadi Muslims "
        "built low, fortified, sometimes half-underground mosques across an island of menzel "
        "farmsteads and rain-cisterns; at El Ghriba (Africa’s oldest synagogue) Jewish pilgrims "
        "gather each spring for Lag BaOmer. UNESCO inscribed the whole island in 2023.",
        "Myth & Religion",
        "UNESCO WHC 2023 (whc.unesco.org/en/list/1932); El Ghriba (Wikipedia); The Arab Weekly",
        "Djerba Jewish community committee; Djerba heritage-management office (INP); island sobas/imams",
        "One island as one fabric; coexistence without postcard; the spring pilgrimage",
        "A fortified Ibadi mosque at prayer, a menzel courtyard, the El Ghriba procession",
        "Consent and discretion at the pilgrimage; access via the community committee",
    ],
    [
        "Amazigh hill-villages — Chenini & Douiret",
        "On the ridges of the Dahar above Tataouine, Chenini and Douiret cling to the rock around "
        "honeycombed ghorfa granaries and a white mosque, and a dwindling number of elders still "
        "speak Tamazight. Families have moved to ‘new’ villages with roads and water, leaving the "
        "old houses to the wind — one of the last living Berber-speaking pockets of a country "
        "that mostly forgot it was Amazigh.",
        "Local communities",
        "Carthage Magazine (Amazigh heritage); WildyNess (Chenini/Douiret); Discover Tunisia",
        "Local Amazigh cultural associations (Tataouine); village elders; the Seven Sleepers mosque keepers",
        "Endangered language; the granary as memory; old village vs new village",
        "Follow a family keeping a hill house and a ghorfa; the olive-and-barley year; Tamazight at home",
        "Identify Tamazight-speaking elders with the association; respect for a thinning generation",
    ],
    [
        "Testour — Andalusian Moriscos & the malouf",
        "Testour, in the Medjerda valley, was rebuilt in the 17th century by Moriscos expelled "
        "from Spain; its Great Mosque carries a clock whose hands turn counter-clockwise — read "
        "locally as the exiles’ longing to wind time back toward Andalusia — and the town keeps "
        "the malouf, the Arab-Andalusian music carried across the sea, alive at Tunisia’s oldest "
        "traditional-music festival.",
        "Arts & Culture",
        "The Arab Weekly (Testour); The National (pomegranate/Andalusian); Tunisian malouf (Wikipedia)",
        "Testour Malouf Festival organisers; local malouf musician families; the Great Mosque custodians",
        "A four-century Mediterranean exile still audible; the backward clock; the brick mosque",
        "Spend the festival with a musician family; follow the Andalusian thread through town and orchard",
        "Festival is the datable scene (summer); confirm the 2026 dates",
    ],
    [
        "The seaside oasis of Gabès & henna",
        "Gabès holds the Mediterranean’s only seaside oasis — a three-tier garden of date palm, "
        "pomegranate, henna and vegetable running down to the shore, watered for millennia — now "
        "poisoned at its edge by the state phosphate works that have piled phosphogypsum into the "
        "gulf for fifty years. Farmers describe henna and pomegranate failing as salt and "
        "pollution rise.",
        "Geopolitics",
        "Med-O-Med (Gabès oasis cultural landscape); Coda Story; TIMEP; SEVENSEAS",
        "Gabès oasis farmers & henna growers; ASOC / local environmental associations; agronomy researchers",
        "A living landscape and the industry pressing on it; henna; the layered garden vs the chimneys",
        "Document the oasis from inside — a henna grower, the irrigation — against the industrial margin",
        "Tension, not indictment; the gardeners are the subject, not the smokestack",
    ],
    [
        "Stambeli — Black Tunisia’s healing trance",
        "In a few houses in Tunis — above all Dar Barnou, once a refuge for freed slaves — the "
        "descendants of sub-Saharan people brought across the Sahara keep Stambeli, a night music "
        "of the gambri lute and iron shqasheq that calls a pantheon of sub-Saharan spirits and "
        "Muslim saints to heal through trance. It survives in only a handful of brotherhoods.",
        "Myth & Religion",
        "Stambeli (Wikipedia); R. Jankowsky, Stambeli (U. Chicago Press); Pan-African Music; The Arab Weekly",
        "Dar Barnou / Sidi Saad Stambeli troupes; ethnomusicologists (Jankowsky network); CMAM (Ennejma Ezzahra)",
        "Black Tunisia’s living liturgy; the gambri and shqasheq; the colours of the spirits; trance",
        "Sit a full Stambeli night with one of the last troupes; the tuning, incense and dance",
        "Consent-led; ritual not spectacle; identity protection where needed",
    ],
    [
        "The Sidi Daoud tuna trap (matanza)",
        "At Sidi Daoud, on the tip of Cap Bon, fishermen set the madrague — a labyrinth of nets "
        "the locals call the matanza — to trap migrating Atlantic bluefin between May and July, a "
        "Phoenician-old technique that ends in a churning chamber of tuna hauled by hand. Quotas, "
        "industrial fleets and warming water have all but stilled one of the Mediterranean’s last "
        "working tuna traps.",
        "Food",
        "INSTM Bulletin (madrague Sidi Daoud, art. 713); Mattanza (Wikipedia); harissa.com",
        "Sidi Daoud fishery & cannery; El Haouaria fishermen; INSTM tuna researchers",
        "A violent, ancient sea-harvest; the net-rooms; the watch for the school; the haul",
        "Embed for a season; build of the trap, the matanza itself, the shore",
        "Seasonal (May–Jul); confirm the trap still runs in 2026 and access",
    ],
    [
        "Cap Bon orange-blossom distillation (neroli/zhar)",
        "Each April and May around Nabeul, families pick the bitter-orange blossom by hand before "
        "dawn, when the scent is strongest, and distil it over slow fire in copper alembics into "
        "zhar — orange-blossom water for sweets and grief and celebration — and the precious "
        "neroli oil. Roughly a million flowers make a kilogram of oil; for two weeks the whole "
        "coast smells of one tree.",
        "Food",
        "Africanews (Nabeul orange blossom); Middle East Eye (floral waters); Biolandes (neroli)",
        "Nabeul/Hammamet distilling families; Cap Bon agricultural cooperatives; ONAT",
        "A Mediterranean perfume-and-kitchen tradition; the pre-dawn picking; the qattara still",
        "Follow one family’s blossom week — picking, distilling, bottling",
        "Seasonal (Apr–May); the whole community works one fortnight",
    ],
    [
        "Mahdia wedding silk (keswa el-kbira)",
        "In the sea-walled medina of Mahdia, weavers still work tall draw-looms — a technology "
        "that crossed the Mediterranean from Iberia, the craft brought by Jewish artisans — to "
        "make narrow silk strips threaded with gold and silver for the keswa el-kbira, the "
        "layered wedding dress that signals a Sahel woman’s married status.",
        "Arts & Culture",
        "Carthage Magazine (Tunisian wedding); Discover Tunisia (regional crafts); shirahime.ch",
        "Mahdia silk-weaving workshops; bride families; the medina artisans’ association",
        "A marriage read in the gold of the cloth; the draw-loom; the dressing of the bride",
        "Portrait a master weaver and a bride’s family through the making and wearing of a keswa",
        "Best filmed around a real wedding, with consent; an ageing craft pressed by imports",
    ],
    [
        "Kairouan — carpets & makroudh",
        "Kairouan, Islam’s fourth holy city and its first in Africa, is also Tunisia’s carpet "
        "capital, where women knot the zerbia — hundreds of thousands of knots a square metre — in "
        "workshops around the Great Mosque, and bake the date-and-semolina makroudh the city "
        "claims as its own. Both loom and oven are pressed by cheaper imports.",
        "History",
        "Discover Tunisia (weaving & carpets); Folk Culture (Kairouan carpet); WildyNess (workshop)",
        "Kairouan carpet-weaving families; ONAT Kairouan; medina makroudh houses",
        "The holy city seen through women’s hands; the knotting frame; the honeyed pastry",
        "Work the medina with one weaving family and one makroudh house",
        "Living; name a workshop and a pastry house on the ground",
    ],
    [
        "Aïn Draham & the Kroumirie cork country",
        "In the Kroumirie range above Tabarka, where it snows in winter and the cork-oak forest is "
        "unlike anywhere else in Tunisia, the Khroumir mountain people harvest cork on its long "
        "bark-cycle, herd and gather among the trees, and keep a green, wet, half-forgotten north "
        "that contradicts the country’s Saharan image.",
        "Local communities",
        "Carthage Magazine (Aïn Draham); Tunisi.info (Kroumirie); Take Your Backpack",
        "Aïn Draham cork-harvesting families; DGF (forestry) Jendouba; Kroumirie forest cooperatives",
        "The Mediterranean Tunisia of oak and rain; the orange-stripped trunks; highland village life",
        "Follow a cork-harvesting family through a summer stripping season and the forest year",
        "Seasonal (summer stripping); a Tunisia of oak bark and snow",
    ],
]

CONTACTS_HEADERS = [
    "Name", "What they do", "Useful for", "Contact", "Status",
    "(To be) Contacted by", "Notes",
]

CONTACTS = [
    [
        "Institut français — Mediterranean photography support",
        "Funder / programme (publication + production aid for Mediterranean photographic work)",
        "The grant itself; confirm strands (one call or two), applicant rules, French-routing, deadlines",
        "IF candidate portal (IF Prog / eMundus, ifprog.emundus.fr)", "To contact", "INLAND / Producer",
        "Deadline 2026-10-01; confirm whether applicant must be photographer or publisher",
    ],
    [
        "Institut français de Tunisie (Tunis)",
        "The IF network office in Tunisia (cultural programming, partnerships)",
        "In-country routing, possible local partner/venue, introductions, a Tunis presentation",
        "[find — IFT Tunis comms]", "To approach", "INLAND / Producer",
        "Likely key for IF Mediterranean routing and a Tunisian launch",
    ],
    [
        "Kerkennah fishing community (charfia)",
        "Charfia palm-fence fishing families on the Kerkennah Islands",
        "Signature story; a named skipper/family; the autumn net-setting and dawn lift",
        "[find — via Sfax/Kerkennah fishers’ networks]", "To approach", "INLAND / partner",
        "UNESCO-documented; identify a named family on scouting",
    ],
    [
        "INSTM — National Institute of Marine Sciences & Technologies (Sfax/Salammbô)",
        "Tunisian marine-science institute; research on charfia and the Sidi Daoud madrague",
        "Charfia & tuna-trap access, route knowledge, the working calendars",
        "[find — INSTM Sfax/Salammbô]", "To approach", "INLAND / Producer",
        "Authors of the Sidi Daoud madrague bulletin; scientific introduction route",
    ],
    [
        "Sejnane women’s pottery associations (Bizerte)",
        "Cooperatives of the women potters of Sejnane",
        "Pottery story; a named maker; the full clay-to-firing cycle",
        "[find — Sejnane cooperative]", "To approach", "INLAND / partner",
        "UNESCO ICH 2018; women-led craft",
    ],
    [
        "Djerba Jewish community committee (El Ghriba)",
        "Manages the El Ghriba synagogue and the Lag BaOmer pilgrimage",
        "El Ghriba access; consent and protocol for the spring pilgrimage",
        "[find — Djerba community committee]", "To approach", "INLAND / partner",
        "Consent and discretion; spring date",
    ],
    [
        "Institut National du Patrimoine (INP) — Djerba site management",
        "National heritage institute; manages the Djerba World Heritage property",
        "Ibadi mosques and menzel access; the island settlement story; official protocol",
        "[find — INP Djerba office]", "To approach", "INLAND / partner",
        "UNESCO WHC 2023 site office",
    ],
    [
        "Amazigh cultural associations (Tataouine / Dahar)",
        "Local associations safeguarding Tamazight language and ksour heritage",
        "Chenini/Douiret story; Tamazight-speaking elders; ghorfa access",
        "[find — Tataouine association]", "To approach", "INLAND / partner",
        "Endangered language; community-led introductions",
    ],
    [
        "Testour Malouf Festival (Béja)",
        "Tunisia’s oldest traditional-music festival; Andalusian malouf",
        "Testour story; malouf musician families; the festival as datable scene",
        "[find — festival organisers]", "To approach", "INLAND / Producer",
        "Confirm 2026 festival dates",
    ],
    [
        "Centre des Musiques Arabes et Méditerranéennes (CMAM, Ennejma Ezzahra, Tunis)",
        "National centre for Arab & Mediterranean music; archives and performers",
        "Stambeli & malouf depth; introductions to troupes; archival sound",
        "[find — CMAM Tunis]", "To approach", "INLAND / Producer",
        "Scholarly/archival route into ritual and Andalusian music",
    ],
    [
        "Dar Barnou / Sidi Saad Stambeli troupe (Tunis)",
        "One of the last Stambeli brotherhoods/houses in Tunis",
        "Stambeli story; a full healing night; consent-led access",
        "[find — via CMAM / Jankowsky network]", "To approach", "INLAND / partner",
        "Consent-sensitive; ritual not spectacle",
    ],
    [
        "Gabès oasis farmers & environmental associations",
        "Henna growers and oasis cultivators; local environmental groups (Gulf of Gabès)",
        "Gabès story; a henna grower; the layered garden; the industrial-edge tension",
        "[find — Gabès oasis cooperative / ASOC]", "To approach", "INLAND / partner",
        "Frame as the gardeners’ story; tension not indictment",
    ],
    [
        "ONAT — Office National de l’Artisanat Tunisien",
        "National crafts office (carpets, ceramics, weaving, distillation)",
        "Kairouan carpets, Mahdia silk, Nabeul distillation, Sejnane pottery — named workshops",
        "[find — ONAT regional offices]", "To approach", "INLAND / Producer",
        "Routes to named artisans across several leads",
    ],
    [
        "Sidi Daoud fishery & cannery (Cap Bon)",
        "The historic tuna trap and cannery at Sidi Daoud",
        "Matanza story; the net-rooms; the May–July season; access at El Haouaria",
        "[find — Sidi Daoud fishery]", "To approach", "INLAND / partner",
        "Confirm the trap still operates in 2026",
    ],
]


def style_sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    for r, row in enumerate(rows, start=2):
        ws.append(row)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 96
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build(path="INLAND_Tunisia_Topics_Contacts.xlsx"):
    wb = Workbook()
    topics = wb.active
    topics.title = "Topics"
    style_sheet(topics, TOPICS_HEADERS, TOPICS,
                widths=[26, 60, 16, 30, 28, 30, 34, 30])
    contacts = wb.create_sheet("Contacts")
    style_sheet(contacts, CONTACTS_HEADERS, CONTACTS,
                widths=[34, 42, 40, 30, 14, 22, 32])
    wb.save(path)
    print(f"Saved {path}: Topics ({len(TOPICS)} rows), Contacts ({len(CONTACTS)} rows)")


if __name__ == "__main__":
    build()
