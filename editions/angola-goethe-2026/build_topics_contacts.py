"""
INLAND Angola — Topics + Contacts spreadsheet builder.
Two sheets (Topics, Contacts) per _inland/deliverables.md. INLAND house style.
Run from this directory: `python3 build_topics_contacts.py`
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# House palette
BG = "F9F4EE"
ACCENT = "400505"
HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
ALT_FILL = PatternFill("solid", fgColor=BG)
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10, color="111111")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D8D2CA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOPICS_HEADERS = [
    "Topic", "Description", "Category", "Links", "Contacts",
    "Interesting for", "Ideas for development", "Notes",
]

TOPICS = [
    [
        "Sona — writing in sand (Lunda-Chokwe)",
        "In the cleaned, dotted sand of the Lunda, elder masters called akwa kuta sona draw a "
        "single unbroken geometric line that holds a proverb, a fable or a piece of cosmology. "
        "It may be a pre-colonial tradition over two thousand years old. UNESCO inscribed it as "
        "Intangible Cultural Heritage in December 2023 — recognised at the moment it is most "
        "fragile, depending on a shrinking number of elders around Saurimo and Dundo.",
        "Arts & Culture",
        "UNESCO ICH 2023 (Ver Angola; Macao News); Lusona (Wikipedia); FLM Journal",
        "Lueji A'Nkonde University; Dundo Museum; Ministry of Culture",
        "The defining cover image; photography of hands and lamplight; the maths-without-maths angle",
        "Follow one master through a night of storytelling; pair drawings with the proverbs they encode",
        "Endangered; identify named masters with the University during scouting",
    ],
    [
        "The ondjango council fire (Ovimbundu)",
        "The ondjango is the Ovimbundu council house, built around a fire kept permanently alight "
        "as a sign of the soba's lineage and continuity. Bailundo (Huambo) is the centre of "
        "Angola's largest ethnic group; King Ekuikui VI, enthroned by the ombala elders in 2021, "
        "remains a living community authority who has travelled to Brazil's quilombo descendants. "
        "Governance by metaphoric speech, not a museum piece.",
        "History",
        "Bailundo (Wikipedia); UNILAB / NJINGA e SEPÉ; Africanews; Ver Angola",
        "Bailundo royal court (King Ekuikui VI's ombala); Huambo cultural authorities",
        "A living institution of speech and consensus; the always-lit fire as the edition's emblem",
        "Sit in on a council night; trace the wood the young bring to feed the fire; the king's authority",
        "Access via the court; respectful protocol with the soba",
    ],
    [
        "Mwila & Mucubal — hair, herd and drought",
        "On the Huíla plateau and in Namibe the Mwila wear ochre nontombi plaits and the Mucubal "
        "the ompota headdress — hairstyles that read as a social language of life-stage, marriage "
        "and mourning. The cattle economy is the base of religious and social identity, and the "
        "ethnoveterinary knowledge of the transhumance routes is documented as endangered under "
        "drought and policy pressure.",
        "Nature",
        "Minority Rights Group; kwekudee tripdown; PMC ethnoveterinary study 2024",
        "Huíla/Namibe pastoral communities; transhumance-route researchers (PMC study authors)",
        "Portraiture of body-as-archive; the herd and the dry-season move; climate pressure",
        "Walk a transhumance leg; document a hairstyle being made and read; herd-healing knowledge",
        "Avoid the tourist 'tribal portrait' cliché; consent and dignity, custodian-led",
    ],
    [
        "Offerings to Kianda — the Festa da Ilha",
        "Every second Friday of November the Axiluanda — Luanda's original fishing people, whose "
        "name is the city's own — cast food and symbolic offerings into the sea to Kianda, the "
        "female water spirit who protects fishermen. A midnight rite that persists as a luxury "
        "waterfront redevelopment pushes the fishers off their own sandbar.",
        "Myth & Religion",
        "Kianda (Wikipedia); Ilha de Luanda (Wikipedia)",
        "Axiluanda fishing community (Ilha de Luanda); Luanda cultural organisations",
        "The signature recurring scene; night-sea offerings; two relationships to the same water",
        "Shoot the full feast; portrait the fishers; pair with the redevelopment squeezing them out",
        "Datable to November; tension not antagonism toward the development",
    ],
    [
        "The mask that dances a woman — mukanda / mwana pwo",
        "The world's museums are full of Chokwe masks; the rite that animates them is barely "
        "filmed. During mukanda, the months-long boys' circumcision-and-education camp, a male "
        "dancer becomes mwana pwo, an idealised female ancestor, entertaining the separated "
        "mothers and teaching Chokwe law and craft. The gap between the vitrine and the dust of "
        "the camp is the story.",
        "Arts & Culture",
        "Chokwe people (Wikipedia); High Museum / DIA mask scholarship",
        "Dundo Museum; Chokwe sobas; Lueji A'Nkonde University",
        "The living dance vs the collected object; the dry-season camps; mothers and sons",
        "Document a mwana pwo performance and a carver; the museum-and-camp diptych",
        "Discretion around the seclusion itself; consent-led; dry season",
    ],
    [
        "The coffee that went wild — Amboim robusta",
        "Angola was the world's third-largest coffee producer until the war turned plantations "
        "into wild bush. Now families harvest from trees their grandparents planted, and the "
        "EU/AFD-financed Mukafe scheme (since 2023) is replanting Amboim's prized robusta in "
        "Kwanza Sul. A story of return, told through one harvest on land the war made forest.",
        "Food",
        "Perfect Daily Grind; STiR Coffee & Tea; Coffee production in Angola (Wikipedia)",
        "Mukafe programme (INCA / AFD); Cafangol (Amboim facility); family cooperatives",
        "Return and revival; the feral roça reclaimed; harvest-season photography",
        "Follow one family through a harvest; the abandoned colonial roça now forest; the replant",
        "~78% family farming; harvest May–Aug; name a family/cooperative in the field",
    ],
    [
        "The council that outlived the king — Mbanza Kongo lumbu",
        "At Mbanza Kongo (Zaire province), UNESCO-listed capital of the old Kongo kingdom, there "
        "has been no king for decades — yet the lumbu, a tribunal of twenty-one elders that "
        "advised the throne for over five hundred years, still meets. An institution that "
        "outlived the institution it served, in a city where the royal past lies underfoot.",
        "History",
        "UNESCO WHC 2017; Met Museum (Kongo)",
        "Mbanza Kongo lumbu elders; UNESCO site-management office",
        "Pure 'what survived'; the elders' faces; the buried royal city and its sacred trees",
        "Attend a council sitting; map the sacred topography; oral memory of the kingdom",
        "Protocol with the elders; the wider royal ritual is much reduced",
    ],
    [
        "Njinga's footprints — Pungo Andongo",
        "At the black monoliths of Pungo Andongo (Malanje) — the Ndongo capital and Queen "
        "Njinga's 17th-century stronghold against the Portuguese — footprints worn into the rock "
        "are venerated locally as the queen's own. National memory made physical and visited; a "
        "400-year-old resistance still touched by hand.",
        "History",
        "Black Rocks at Pungo Andongo (Wikipedia); Atlas Obscura",
        "Pungo Andongo local custodians/guides; Malanje cultural authorities",
        "A shrine without a temple; landscape-as-archive; informal pilgrimage",
        "Shoot the footprints as a place of veneration; the rocks as refuge; pair with Kalandula circuit",
        "Pairs with the Malanje route; light at dawn/dusk on the monoliths",
    ],
    [
        "Those who ate the ox — efundula",
        "The efundula (efico/efuko) female initiation marks a girl's passage to womanhood among "
        "the Kwanyama and Handa of Cunene — historically including the symbolic 'eating of the "
        "ox' — surviving in tension with the school and the state, as young mothers are caught "
        "between the rite and school dropout. Continuity and its costs, held without judgement.",
        "Myth & Religion",
        "Efundula research (Academia.edu); UCP study, Cunene; ResearchGate (Handa efuko)",
        "Cunene community leaders; UCP / academic researchers on efundula",
        "Continuity vs schooling; a women's rite; the negotiation with modern law",
        "Follow the cycle with the community's lead; the feast; voices of young mothers",
        "Consent-sensitive; identities protected where needed; never staged",
    ],
    [
        "Salt and the long boats — Baía Farta",
        "In one Benguela bay, two squeezed coastal trades: solar salt evaporated in the pans that "
        "named the beach, and an artisanal fishing economy of tens of thousands, both pressed by "
        "industrial trawling and packaged substitutes. The food infrastructure of the coast, "
        "still made by hand at the Casa dos Pescadores Artesanais.",
        "Food",
        "Journey Gourmet (Benguela); trackstick (Baía Farta); DLIST Benguela",
        "Casa dos Pescadores Artesanais de Baía Farta; Benguela fishing cooperatives",
        "Hand-made food infrastructure; salt and sea on one shore; the artisanal-vs-industrial squeeze",
        "Shoot a salt-raking day and a fishing dawn; the drying racks; the Casa dos Pescadores",
        "Working calendar; economic pressure from trawlers",
    ],
    [
        "Clay and fibre — women potters and weavers",
        "Coil pottery and sisal/palm basketry, made by methods passed down generations across the "
        "highlands and east (Huambo, Bié, Moxico), now half-commercialised for an urban and "
        "tourist market — the domestic vessels of Angolan food-and-household life, and the women "
        "who still make them.",
        "Arts & Culture",
        "African Sahara; Design Milk (Angolan basketry & ceramics)",
        "Highland/eastern craft workshops (to be named); artisan cooperatives",
        "Women's craft economy; hand and material; the market as lifeline and pressure",
        "Portrait one named maker through a full firing/weaving; the market stall",
        "Needs a named workshop/maker in the field to clear the bar",
    ],
    [
        "The most expensive city and the musseque",
        "Luanda's oil-built towers stand a few hundred metres from self-built musseques and the "
        "kitanda markets that feed the city. Anchored on one human — a fisherman from the Ilha, a "
        "market woman — it shows two relationships to the same sea and land coexisting, as "
        "portrait rather than abstraction.",
        "Geopolitics",
        "Ilha de Luanda (Wikipedia); INLAND field framing",
        "Ilha de Luanda fishers (links to Kianda); kitanda market vendors",
        "The petro-city's underside; one human anchor; two economies, one shore",
        "Follow one Ilha fisher between the sandbar and the skyline; the kitanda day",
        "Needs a human anchor; links to lead 04 (Kianda)",
    ],
]

CONTACTS_HEADERS = [
    "Name", "What they do", "Useful for", "Contact", "Status",
    "(To be) Contacted by", "Notes",
]

CONTACTS = [
    [
        "Lueji A'Nkonde University (Lunda Norte)",
        "Angolan public university; coordinated the UNESCO sona ICH nomination (2023) with the Ministry of Culture",
        "Lead African-partner candidate; sona & Chokwe access, consent, co-creation, Angola presentation",
        "[find — university comms office]", "To approach", "INLAND / Producer",
        "Documented heritage-safeguarding record; strongest sourced partner fit",
    ],
    [
        "Dundo Museum (Lunda Norte)",
        "Ethnographic museum holding major Chokwe collections (masks, material culture)",
        "Mukanda/mwana pwo depth; candidate exhibition venue; introductions to sobas",
        "[find]", "To approach", "INLAND / partner", "Pair living rite with collected object",
    ],
    [
        "Bailundo royal court — King Ekuikui VI (Tchongolola Tchongonga)",
        "Reigning Ovimbundu monarch of Bailundo, enthroned by the ombala council 2021",
        "Ondjango stories; access and protocol in the Huambo highlands",
        "[find — via Huambo cultural authorities]", "To approach", "Angolan partner",
        "Respectful protocol with the soba required",
    ],
    [
        "Geração 80 (Luanda)",
        "Independent Angolan film / visual-culture collective",
        "Editorial co-creation candidate; Luanda exhibition leg; urban anchor story",
        "[confirm interest]", "To approach", "INLAND / Producer",
        "Confirm current activity and interest before naming in application",
    ],
    [
        "National Museum of Anthropology (Luanda)",
        "National ethnographic museum",
        "Ethnographic depth; candidate Angola exhibition venue",
        "[find]", "To approach", "INLAND / partner", "Capital-city presentation route",
    ],
    [
        "Mukafe programme (INCA / AFD-financed)",
        "EU/AFD-financed coffee-revival scheme in Uíge, Kwanza Sul, Kwanza Norte (since 2023)",
        "Amboim robusta story; routes to coffee families and washing stations",
        "[find — INCA / AFD Angola]", "To approach", "INLAND / partner",
        "Confirm a specific family/cooperative on the ground",
    ],
    [
        "Casa dos Pescadores Artesanais de Baía Farta (Benguela)",
        "Artisanal-fishing centre teaching local techniques; salt-pan area nearby",
        "Salt & fishing story; access to fishers and the working calendar",
        "[find]", "To approach", "INLAND / partner", "Coastal cluster anchor",
    ],
    [
        "Mbanza Kongo UNESCO site-management office (Zaire)",
        "Manages the Mbanza Kongo World Heritage Site",
        "Access to the lumbu elders; the sacred topography of the buried capital",
        "[find]", "To approach", "INLAND / partner", "Protocol with the council of elders",
    ],
    [
        "PMC ethnoveterinary-transhumance researchers (2024 study)",
        "Academics documenting endangered herd-healing knowledge on SW Angola transhumance routes",
        "Mwila/Mucubal story; introductions and route knowledge",
        "[via the published paper, PMC10933900]", "To approach", "INLAND / Producer",
        "Academic introduction route into pastoral communities",
    ],
    [
        "UCP / efundula researchers (Cunene)",
        "Researchers on efundula and school dropout among young mothers in Cunene",
        "efundula story; consent-first community introductions",
        "[via the published study]", "To approach", "INLAND / Producer",
        "Consent-sensitive; use academic introductions",
    ],
    [
        "Goethe-Institut — Africa–Europe Partnerships for Culture",
        "Funder / programme (Connect & Create)",
        "The grant itself; confirm strand rules, ceiling, co-financing, prior-contact requirement",
        "Goethe Application Portal (GAP)", "To contact", "INLAND / Producer",
        "Confirm Visual Arts ceiling and mobility windows before submission",
    ],
]


def style_sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    for r, row in enumerate(rows, start=2):
        ws.append(row)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 96
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build(path="INLAND_Angola_Topics_Contacts.xlsx"):
    wb = Workbook()
    topics = wb.active
    topics.title = "Topics"
    style_sheet(topics, TOPICS_HEADERS, TOPICS,
                widths=[26, 60, 16, 30, 28, 30, 34, 30])
    contacts = wb.create_sheet("Contacts")
    style_sheet(contacts, CONTACTS_HEADERS, CONTACTS,
                widths=[30, 42, 40, 26, 14, 22, 34])
    wb.save(path)
    print(f"Saved {path}: Topics ({len(TOPICS)} rows), Contacts ({len(CONTACTS)} rows)")


if __name__ == "__main__":
    build()
