"""
INLAND Georgia — Topics + Contacts spreadsheet builder.
Two sheets (Topics, Contacts) per _inland/deliverables.md. INLAND house style.
For the W. Eugene Smith Grant in Humanistic Photography 2026.
Run from this directory: `python3 build_topics_contacts.py`
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# House palette
BG = "F9F4EE"
ACCENT = "400505"
HEAD_FILL = PatternFill("solid", fgColor=ACCENT)
ALT_FILL = PatternFill("solid", fgColor=BG)
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10, color="111111")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D8D2CA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TOPICS_HEADERS = [
    "Topic", "Description", "Category", "Links", "Contacts",
    "Interesting for", "Ideas for development", "Notes",
]

TOPICS = [
    [
        "Tusheti transhumance (the Abano Pass)",
        "Around a tenth of Tusheti’s people still live to the rhythm of transhumance, driving "
        "sheep — and, ever more rarely, riding the migration on horseback — between the Kakheti "
        "lowlands and the high summer grass of Tusheti, climbing and descending the Abano Pass at "
        "2,850 m, one of the most dangerous roads in the country. National Geographic followed a "
        "2019 crossing with shepherd Giorgi Karsamauli; RFE/RL profiled a woman keeping the mounted "
        "horse-migration alive as most switch to trucks. A half-millennium-old movement, on a real date.",
        "Nature",
        "Atlas Obscura (transhumance Tusheti); National Geographic (Caucasus migration); RFE/RL (horse migration); Milk Trekker",
        "Giorgi Karsamauli (shepherd); the Tusheti horsewoman (RFE/RL); Tusheti Protected Areas",
        "The project’s signature recurring image — the pass in cloud, the night camps, the dogs",
        "Follow one family through the spring drive up and the autumn descent",
        "Seasonal (late May / Sept); confirm 2026 dates and whether the mounted ride runs; consent",
    ],
    [
        "Ushguli tower-keeping families",
        "Ushguli, at the confluence of the Inguri and Black rivers, is among Europe’s highest "
        "inhabited places — a UNESCO site of 200-plus medieval koshki defence towers that were at "
        "once home, fortress and treasury. A handful of families still live the old way among them: "
        "Roland and Gulo Chelidze keep cattle, pigs, goats and horses and look after the family’s "
        "tower; horses remain reliable transport across largely roadless high country.",
        "History",
        "UNESCO WHC 709 (Upper Svaneti); National Geographic (Svanetia); Atlas Obscura (Towers of Svaneti); Museum of Wander",
        "Roland & Gulo Chelidze (Ushguli); Svaneti Museum (Mestia); local guesthouse networks",
        "A house, a fortress and a treasury in the same stone; the winter day below the keep",
        "Spend a winter day with one family between byre and tower",
        "Accessible late spring–autumn; winter weather-dependent; re-contact the family",
    ],
    [
        "Svan animist hymns — Lile & Riho",
        "Some of the oldest layers of Georgian polyphony are Svan: hymns like Lile (the ancient Svan "
        "word for the Sun God) and Kviria carry largely pre-Christian texts Svans now often re-read "
        "through a Christian lens. The Mestia ensemble Riho kept these songs for over fifty years under "
        "Islam Pilpani, a ‘Benefactor of Georgian Folk Song’; his son Vakhtang Pilpani now directs it.",
        "Arts & Culture",
        "Georgian Chant (Svan chant; Kviria); Chanting.ge (Svan folk songs); Discogs (Riho); Tsutisopeli",
        "Riho Ensemble; Vakhtang Pilpani (director); Georgian Chanting Foundation",
        "Three braided male voices in a sun-hymn older than the church; a Mestia front-room rehearsal",
        "Sit a rehearsal with Riho; follow the sun-hymn from animist text to Christian re-reading",
        "Year-round rehearsals; consent to photograph; distinct from the language lead",
    ],
    [
        "Gurian krimanchuli (the highest voice)",
        "On the Black Sea side, Gurian polyphony is the most acrobatic of all Georgian song: above the "
        "texture flies the krimanchuli, a high yodel-falsetto of tension and release. Georgian polyphony "
        "was proclaimed by UNESCO in 2001 and inscribed in 2008. In Lanchkhuti, the Turkia family has "
        "been central to keeping and teaching the Gurian repertoire across generations.",
        "Arts & Culture",
        "UNESCO ICH 00008 (Georgian polyphonic singing); chortownia.org (Georgian polyphony); harmonychroniclesmag.com",
        "The Turkia family (Lanchkhuti); the International Centre for Georgian Folk Song (Tbilisi)",
        "The voice that goes up where no other follows; a supra table breaking into krimanchuli",
        "Work a real feast where a three-part song breaks into the high voice; the teaching of it",
        "Year-round; tie to an actual supra; name the singers and confirm consent",
    ],
    [
        "Minankari cloisonné enamel revival",
        "Minankari, Georgian cloisonné enamel, reached its height between the 10th and 15th centuries on "
        "icons and crosses, then was abandoned for centuries. Over the last two decades Tbilisi artists "
        "have revived it by an almost unchanged hand process, centred on the ‘Ornament’ gallery (founded "
        "2001); masters such as Thea Gurgenidze have built names teaching and working the craft.",
        "Arts & Culture",
        "NTD (cloisonné enamel revived in Georgia); mygeotrip (Minankari; Ornament gallery); georgia.to (enamel studios)",
        "Thea Gurgenidze (master); ‘Ornament’ gallery (Tbilisi); allied enamel studios",
        "A medieval church technique alive in a living hand; gold wire and fired colour at the bench",
        "Portrait a master laying wire and firing colour; the master–apprentice relationship",
        "Year-round; confirm a named master and consent",
    ],
    [
        "Qvevri wine & the marani at harvest",
        "Georgia’s 8,000 years of winemaking are lived, not marketed, in the marani — the cellar where "
        "egg-shaped qvevri are buried to the neck and the wine ferments on its skins through the cold "
        "without machinery. UNESCO inscribed the method in 2013; PGI status followed in 2021. The qvevri "
        "themselves are still thrown by a few families at Vardisubani (Kakheti) and Shrosha (Imereti).",
        "Food",
        "UNESCO ICH 00870 (qvevri); wine.gov.ge (National Wine Agency); Decanter (orange wine); winesgeorgia.com",
        "A Kakhetian winemaking family (to name on the ground); Vardisubani qvevri potters; National Wine Agency",
        "The wine made under the floor of the house; the potter firing a vessel taller than a man",
        "Follow a family through rtveli — pressing, sealing the qvevri, the feast — and the qvevri potter",
        "Seasonal (rtveli, late Sept); name the family and the potter",
    ],
    [
        "Tushetian guda cheese & nabadi felt",
        "Tushetian guda is made by herders during the summer transhumance from whole raw cow’s and "
        "sheep’s milk, matured inside a guda — a whole sheepskin sack — and covered with the felted "
        "nabadi burka so it takes the salt evenly. The women run the dairy and the wool: pardagi felt "
        "carpets, the burka cloak, headwear. The EU-backed Alaznistavi cooperative has revived guda.",
        "Food",
        "EU4Georgia (Alaznistavi guda cooperative); Milk Trekker (Tusheti / guda); Wikipedia (Tushetians); nlinemedia",
        "Alaznistavi cooperative (Tusheti); named guda maker (to confirm); Tusheti women wool-workers",
        "A cheese born in a sheepskin on a summer mountain; the nabadi over the curing room",
        "Document a summer dairy hut — milk into skin, the felt, the women’s wool work",
        "Seasonal (Jun–Sept); pairs with the transhumance lead; name a maker via Alaznistavi",
    ],
    [
        "Khevsureti shrine cult (Atengenoba)",
        "In Khevsureti people worship at jvari / khati — stone shrines two metres tall, hung with ibex "
        "horns and a bell — in a syncretism where, behind the Christian names and candles, the rite is "
        "essentially pre-Christian. The shrine is served by an elected khevisberi, a dasturi, and, in "
        "Pirikita Khevsureti, a khatis diasakhlisi (hostess of the shrine). Atengenoba falls 100 days "
        "after Easter, with beer brewed at the shrine and animals offered.",
        "Myth & Religion",
        "Wikipedia (Khevsureti); academia.edu (‘Khatis Diasakhlisi’); Pshav-Khevsureti sacred places (khconsult); georgia-tours.eu",
        "A named khevisberi; a khatis diasakhlisi; the Shatili/Mutso communities; ethnographers of the highlands",
        "A faith older than the church behind the candles; the horned shrine on Atengenoba",
        "Work Atengenoba at a khati shrine — the priest, the beer, the offering — from Shatili/Mutso",
        "Seasonal (Atengenoba ~mid-Jul 2026); consent-led; ritual not spectacle; confirm dates/access",
    ],
    [
        "Pankisi Kist women’s Sufi zikr",
        "The Kist of Pankisi descend from Chechens who crossed the mountains in the 1800s and follow the "
        "pacifist Sufi path of Kunta-Hajji. Pankisi is the only place in the South Caucasus where women "
        "lead the zikr in the mosque — every Friday at the Old Mosque in Duisi, circling and chanting "
        "nazms in the Kist tongue and calling ‘marshua kavkaz’, peace in the Caucasus. Led by elders, "
        "the rite is edging toward silence as younger women turn to a stricter Sunni practice.",
        "Myth & Religion",
        "OC Media (the last women performing zikr in Pankisi); Al Jazeera America (female Sufi mystics); pankisivalley.com",
        "The Duisi women’s zikr circle; the Aznash Laaman ensemble; the Pankisi Valley tourism/heritage association",
        "The only women’s zikr in the South Caucasus; the turning chant; an endangered rite",
        "Photograph the Friday women’s circle with consent; the nazms; the call for peace",
        "Weekly (Fridays); consent-sensitive; protect identities; endangered",
    ],
    [
        "David Gareja — desert monastery on the line",
        "David Gareja is a complex of 19 medieval cave monasteries — some 5,000 cells cut into the "
        "Kakheti semi-desert in the 6th century by St David, one of the Thirteen Assyrian Fathers — with "
        "8th–13th-century frescoes including a rare portrait of Queen Tamar at Udabno. The Soviet-drawn "
        "Georgia–Azerbaijan border runs straight through the complex, making part of it technically "
        "Azerbaijani and the line a quiet, live dispute.",
        "Geopolitics",
        "UNESCO tentative 5224 (David Gareji); Advantour; Atlas Obscura (Davit Gareja); wander-lush.org (2026 access)",
        "David Gareja monastery (Lavra); Georgian Patriarchate; Agency for Cultural Heritage Preservation",
        "Faith, frescoes and a frontier in one frame; a monk on the ridge where the line meets the caves",
        "Frame monastic life against the disputed line; the Queen Tamar fresco at Udabno",
        "Conditional — verify 2026 border access before scheduling Udabno; tension not attack",
    ],
    [
        "Svan — a language dying at home, sung by the young",
        "Svan, an unwritten Kartvelian language perhaps three millennia diverged from Georgian, is "
        "classed as endangered, spoken by a shrinking number of elders even as Svan music enjoys a "
        "youth-led revival driven from Mestia (the Lileo song-and-dance ensemble; young singers learning "
        "the old hymns). The paradox is the story: the songs return to young mouths while the everyday "
        "language thins.",
        "Local communities",
        "Chai Khana (Svans, gatekeepers of the mountains); georgianchant.org (Svaneti revival); UNESCO Atlas of endangered languages",
        "The Lileo ensemble (Mestia); young Svan singers; Svan language activists",
        "Songs returning to young mouths while the language thins; words kept like seed-grain",
        "Follow young singers learning hymns in a tongue fewer of them speak at home",
        "Year-round; distinct from the Riho/Lile lead (this is language & generation)",
    ],
    [
        "Georgian Jews of Oni & Kutaisi",
        "Jews have lived in Georgia for some 26 centuries, in what is often called the one country where "
        "they were never persecuted as Jews. That long presence is now very thin. The Oni synagogue "
        "(1895, Moorish Revival), the oldest functioning synagogue in Georgia, once held over a thousand; "
        "today fewer than 20 worship there. Kutaisi’s Jewish quarter on Boris Gaponov Street, once 30,000 "
        "strong, counts fewer than a hundred families across its 19th-century synagogues.",
        "Local communities",
        "Wikipedia (Oni Synagogue; Kutaisi Synagogue); Gil Travel (2,600 years of Jewish history); jguideeurope.org (Oni)",
        "The Oni Jewish community; the Kutaisi Jewish community; cultural-routes.gov.ge (Jewish heritage route)",
        "One of the world’s oldest diasporas down to a last minyan; a Sabbath in a hall built for a thousand",
        "Document a Sabbath where the minyan is barely made; the keepers of the old synagogues",
        "Year-round (confirm festival dates); discretion; endangered community",
    ],
]

CONTACTS_HEADERS = [
    "Name", "What they do", "Useful for", "Contact", "Status",
    "(To be) Contacted by", "Notes",
]

CONTACTS = [
    [
        "W. Eugene Smith Memorial Fund",
        "Funder — annual grant in humanistic photography ($30,000 main + two $10,000 finalist grants)",
        "The grant itself; confirm 2026 cycle dates, submission portal, and that the applicant is the photographer",
        "smithfund.org/eugene-smith-grant (online submission portal)", "To confirm", "INLAND / Producer",
        "2025 cycle: deadline ~7 Oct, $50 fee, fee waivers via partner orgs; call opens ~1 Jul. Confirm 2026.",
    ],
    [
        "Commissioned INLAND photographer (named applicant)",
        "The individual documentary photographer who is the grant applicant on the Georgia body of work",
        "The portfolio (past work), the proposed-project edit, the bio/CV — the application is in their name",
        "[INLAND to confirm which photographer]", "To confirm", "INLAND",
        "Smith applicant must be an individual; confirm the Issue 02 Georgia work is theirs to present",
    ],
    [
        "Giorgi Karsamauli (Tusheti shepherd)",
        "Shepherd documented on the Abano-Pass migration (National Geographic, 2019)",
        "Lead 01 — the transhumance; a named protagonist for the signature image",
        "[find — via Tusheti shepherd networks]", "To approach", "INLAND / partner",
        "Re-contact; confirm he rides/walks the 2026 season and consents",
    ],
    [
        "Tusheti horse-migration custodian (RFE/RL)",
        "Woman keeping the mounted horse-migration alive as others switch to trucks",
        "Lead 01 — the rare mounted migration; a strong single protagonist",
        "[find — via RFE/RL Georgia / Tusheti PA]", "To approach", "INLAND / partner",
        "Identify by name; verify 2026 ride; may merge into the transhumance story",
    ],
    [
        "Roland & Gulo Chelidze (Ushguli)",
        "Tower-keeping family in Upper Svaneti — cattle, horses and the family defence tower",
        "Lead 02 — the tower and the byre; a named family among the koshki",
        "[find — via Ushguli guesthouse networks / Svaneti Museum]", "To approach", "INLAND / partner",
        "Re-contact and confirm; winter access weather-dependent",
    ],
    [
        "Riho Ensemble — Vakhtang Pilpani (Mestia)",
        "Director of the Mestia regional folk ensemble; keepers of Svan sun-hymns",
        "Lead 03 — Lile / Kviria; a rehearsal of the animist hymns",
        "[find — via Georgian Chanting Foundation]", "To approach", "INLAND / Producer",
        "Son of the late Islam Pilpani; year-round rehearsals",
    ],
    [
        "The Turkia family (Lanchkhuti, Guria)",
        "Family central to the Gurian polyphony repertoire and the krimanchuli",
        "Lead 04 — the high voice at a supra; teaching across generations",
        "[find — via Gurian folk networks / Tbilisi folk-song centre]", "To approach", "INLAND / partner",
        "Tie to a real feast; name the singers; confirm consent",
    ],
    [
        "Thea Gurgenidze / ‘Ornament’ gallery (Tbilisi)",
        "Minankari master and the dedicated cloisonné-enamel revival gallery (founded 2001)",
        "Lead 05 — the enamel; bench work and the master–apprentice relationship",
        "[find — Ornament gallery, Tbilisi]", "To approach", "INLAND / Producer",
        "Confirm a named master and consent; year-round",
    ],
    [
        "Vardisubani qvevri potters (Kakheti)",
        "Specialist family potters who throw and fire the buried wine qvevri",
        "Lead 06 — the qvevri made by hand; a vessel taller than a man",
        "[find — Vardisubani, Telavi district]", "To approach", "INLAND / partner",
        "Pair with a Kakhetian winemaking family at rtveli",
    ],
    [
        "Kakhetian winemaking family (marani)",
        "A family making qvevri wine in the traditional marani at harvest",
        "Lead 06 — rtveli; the wine made under the floor of the house",
        "[find — Telavi / Kakheti, via National Wine Agency]", "To approach", "INLAND / partner",
        "Name on the ground; seasonal (late Sept)",
    ],
    [
        "Alaznistavi guda cooperative (Tusheti)",
        "EU-supported cooperative reviving Tushetian guda cheese; women in the dairy",
        "Lead 07 — guda and nabadi felt; a named maker in the summer dairy",
        "[find — via EU4Georgia / Tusheti development networks]", "To approach", "INLAND / partner",
        "Seasonal (summer); pairs with the transhumance lead",
    ],
    [
        "Khevsureti shrine community (khevisberi)",
        "Elected shrine-priest (khevisberi), dasturi and khatis diasakhlisi of a khati shrine",
        "Lead 08 — Atengenoba; the pre-Christian rite at the horned shrine",
        "[find — via Shatili/Mutso communities & highland ethnographers]", "To approach", "INLAND / partner",
        "Consent-led; confirm Atengenoba 2026 dates and access; ritual not spectacle",
    ],
    [
        "Duisi women’s zikr circle / Aznash Laaman (Pankisi)",
        "The Friday women’s Sufi zikr at the Old Mosque, Duisi; the all-female ensemble",
        "Lead 09 — the only women’s zikr in the South Caucasus; an endangered rite",
        "[find — via Pankisi Valley Tourism & Development Assoc.]", "To approach", "INLAND / partner",
        "Consent-sensitive; protect identities; weekly (Fridays)",
    ],
    [
        "David Gareja monastery (Lavra) & Cultural Heritage Agency",
        "The working cave monastery and the national heritage body for the site",
        "Lead 10 — monastic life on the disputed line; the Udabno frescoes",
        "[find — Georgian Patriarchate; cultural-heritage agency]", "To approach", "INLAND / Producer",
        "Verify 2026 border access before scheduling Udabno; tension not attack",
    ],
    [
        "Lileo ensemble & Svan language activists (Mestia)",
        "Youth song-and-dance ensemble and activists for the endangered Svan language",
        "Lead 11 — songs returning to the young while the language thins",
        "[find — Mestia cultural networks]", "To approach", "INLAND / partner",
        "Year-round; distinct from the Riho lead",
    ],
    [
        "Oni & Kutaisi Jewish communities",
        "The keepers of the Oni synagogue (1895) and the Kutaisi Jewish quarter synagogues",
        "Lead 12 — a Sabbath in a hall built for a thousand; one of the oldest diasporas",
        "[find — via community leaders; cultural-routes.gov.ge]", "To approach", "INLAND / partner",
        "Discretion; confirm festival dates; endangered community",
    ],
]


def style_sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    for r, row in enumerate(rows, start=2):
        ws.append(row)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = ALT_FILL
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 110
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build(path="INLAND_Georgia_Topics_Contacts.xlsx"):
    wb = Workbook()
    topics = wb.active
    topics.title = "Topics"
    style_sheet(topics, TOPICS_HEADERS, TOPICS,
                widths=[28, 62, 16, 32, 30, 30, 34, 32])
    contacts = wb.create_sheet("Contacts")
    style_sheet(contacts, CONTACTS_HEADERS, CONTACTS,
                widths=[34, 42, 40, 32, 14, 22, 34])
    wb.save(path)
    print(f"Saved {path}: Topics ({len(TOPICS)} rows), Contacts ({len(CONTACTS)} rows)")


if __name__ == "__main__":
    build()
